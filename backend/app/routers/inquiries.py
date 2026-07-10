from typing import Optional
from datetime import datetime
from io import BytesIO

from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import Inquiry, get_db

router = APIRouter()


class InquiryCreate(BaseModel):
    company: str = ""
    name: str
    position: str = ""
    phone: str
    message: str = ""


class InquiryResponse(BaseModel):
    id: int
    company: str
    name: str
    position: str
    phone: str
    message: str
    is_read: bool
    is_replied: bool
    created_at: datetime
    note: str

    class Config:
        from_attributes = True


class InquiryListResponse(BaseModel):
    total: int
    unread: int
    items: list[InquiryResponse]


class InquiryNoteUpdate(BaseModel):
    note: str


class BatchIdsRequest(BaseModel):
    ids: list[int]


SORT_FIELDS = {
    "created_at": Inquiry.created_at,
    "company": Inquiry.company,
    "name": Inquiry.name,
    "position": Inquiry.position,
}


@router.post("", response_model=InquiryResponse)
async def create_inquiry(
    data: InquiryCreate,
    db: AsyncSession = Depends(get_db),
):
    inquiry = Inquiry(
        company=data.company or "",
        name=data.name,
        position=data.position or "",
        phone=data.phone,
        message=data.message or "",
    )
    db.add(inquiry)
    await db.commit()
    await db.refresh(inquiry)

    try:
        from app.services.notification import send_inquiry_notification
        await send_inquiry_notification(inquiry)
    except Exception:
        pass

    return inquiry


@router.get("", response_model=InquiryListResponse)
async def list_inquiries(
    page: int = 1,
    page_size: int = 20,
    is_read: Optional[bool] = None,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "desc",
    db: AsyncSession = Depends(get_db),
):
    query = select(Inquiry)
    count_query = select(func.count(Inquiry.id))

    if search:
        term = f"%{search}%"
        search_filter = or_(
            Inquiry.company.ilike(term),
            Inquiry.name.ilike(term),
            Inquiry.position.ilike(term),
            Inquiry.phone.ilike(term),
            Inquiry.message.ilike(term),
        )
        query = query.where(search_filter)
        count_query = count_query.where(search_filter)

    if is_read is not None:
        query = query.where(Inquiry.is_read == is_read)
        count_query = count_query.where(Inquiry.is_read == is_read)

    sort_col = SORT_FIELDS.get(sort_by, Inquiry.created_at)
    if sort_order == "asc":
        query = query.order_by(sort_col.asc())
    else:
        query = query.order_by(sort_col.desc())

    if sort_by and sort_by != "created_at":
        query = query.order_by(Inquiry.created_at.desc())

    total_result = await db.execute(count_query)
    total = total_result.scalar()

    unread_result = await db.execute(
        select(func.count(Inquiry.id)).where(Inquiry.is_read == False)
    )
    unread = unread_result.scalar()

    offset = (page - 1) * page_size
    query = query.offset(offset).limit(page_size)

    result = await db.execute(query)
    items = result.scalars().all()

    return InquiryListResponse(total=total, unread=unread, items=items)


@router.get("/export")
async def export_inquiries(
    search: Optional[str] = None,
    is_read: Optional[bool] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "desc",
    db: AsyncSession = Depends(get_db),
):
    query = select(Inquiry)

    if search:
        term = f"%{search}%"
        query = query.where(or_(
            Inquiry.company.ilike(term),
            Inquiry.name.ilike(term),
            Inquiry.position.ilike(term),
            Inquiry.phone.ilike(term),
            Inquiry.message.ilike(term),
        ))

    if is_read is not None:
        query = query.where(Inquiry.is_read == is_read)

    sort_col = SORT_FIELDS.get(sort_by, Inquiry.created_at)
    if sort_order == "asc":
        query = query.order_by(sort_col.asc())
    else:
        query = query.order_by(sort_col.desc())

    result = await db.execute(query)
    items = result.scalars().all()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "联系列表"

        headers = ["序号", "公司名称", "联系人", "招聘岗位", "电话", "备注", "状态", "提交时间", "管理员备注"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="5B8C5A", end_color="5B8C5A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for idx, item in enumerate(items, 1):
            status = "已回复" if item.is_replied else ("已读" if item.is_read else "未读")
            row_data = [
                idx,
                item.company or "",
                item.name,
                item.position or "",
                item.phone,
                item.message or "",
                status,
                item.created_at.strftime("%Y-%m-%d %H:%M:%S") if item.created_at else "",
                item.note or "",
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 30

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"inquiries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="未安装 openpyxl 库，无法导出 Excel")


@router.get("/{inquiry_id}", response_model=InquiryResponse)
async def get_inquiry(inquiry_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Inquiry).where(Inquiry.id == inquiry_id))
    inquiry = result.scalar_one_or_none()
    if not inquiry:
        raise HTTPException(status_code=404, detail="记录不存在")
    if not inquiry.is_read:
        inquiry.is_read = True
        await db.commit()
        await db.refresh(inquiry)
    return inquiry


@router.put("/{inquiry_id}/read")
async def mark_as_read(inquiry_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Inquiry).where(Inquiry.id == inquiry_id))
    inquiry = result.scalar_one_or_none()
    if not inquiry:
        raise HTTPException(status_code=404, detail="记录不存在")
    inquiry.is_read = True
    await db.commit()
    return {"success": True}


@router.put("/{inquiry_id}/reply")
async def mark_as_replied(inquiry_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Inquiry).where(Inquiry.id == inquiry_id))
    inquiry = result.scalar_one_or_none()
    if not inquiry:
        raise HTTPException(status_code=404, detail="记录不存在")
    inquiry.is_replied = True
    await db.commit()
    return {"success": True}


@router.put("/{inquiry_id}/note")
async def update_note(inquiry_id: int, data: InquiryNoteUpdate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Inquiry).where(Inquiry.id == inquiry_id))
    inquiry = result.scalar_one_or_none()
    if not inquiry:
        raise HTTPException(status_code=404, detail="记录不存在")
    inquiry.note = data.note
    await db.commit()
    return {"success": True}


@router.delete("/{inquiry_id}")
async def delete_inquiry(inquiry_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Inquiry).where(Inquiry.id == inquiry_id))
    inquiry = result.scalar_one_or_none()
    if not inquiry:
        raise HTTPException(status_code=404, detail="记录不存在")
    await db.delete(inquiry)
    await db.commit()
    return {"success": True}


@router.put("/batch/read")
async def batch_mark_read(data: BatchIdsRequest, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Inquiry).where(Inquiry.id.in_(data.ids)))
    inquiries = result.scalars().all()
    for inquiry in inquiries:
        inquiry.is_read = True
    await db.commit()
    return {"success": True, "count": len(inquiries)}


@router.put("/batch/reply")
async def batch_mark_replied(data: BatchIdsRequest, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Inquiry).where(Inquiry.id.in_(data.ids)))
    inquiries = result.scalars().all()
    for inquiry in inquiries:
        inquiry.is_replied = True
    await db.commit()
    return {"success": True, "count": len(inquiries)}


@router.delete("/batch")
async def batch_delete(data: BatchIdsRequest, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Inquiry).where(Inquiry.id.in_(data.ids)))
    inquiries = result.scalars().all()
    for inquiry in inquiries:
        await db.delete(inquiry)
    await db.commit()
    return {"success": True, "count": len(inquiries)}
